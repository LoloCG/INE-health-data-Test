'''
Generates a list of maps, where the map structure for each variable is:
[{
    'Variable': 'CCAA', 
    'Diccionario de la variable': 'TCCAA', 
    'Longitud': 2, 
    'Tipo': 'A', 
    'Decimales': None, 
    'Posición': 1, 
    'Orden': 1, 
    'Diccionario ubicado en la hoja…': 'Tablas1', 
    'Descripción': 'Comunidad Autónoma de residencia', 
    'Observaciones': None, 
    'Grupo': 'DATOS DE IDENTIFICACIÓN ',
    'value_labels':{
        '01':'Andalucía',
        '02':'...',
    }
},{
}]

This causes duplication of variable "value_labels".
Codebook contains 432 variables. Even with repeated number of value labels, it does not request the usage of chunking or
other optimization methods.

Positive trait of this duplication is that it avoids lookups into multiple sources.
It could later be normalized by separating into different files for cross referencing or through insertion into SQL with duplication removal.

'''
from pathlib import Path
import json
from openpyxl import load_workbook, worksheet
import logging

# wb = Workbook()

# Path for the original raw catalogue supplied from INE
RAW_ESdE_CATALOG_PATH = Path(r"data\raw\datos_2023\ESdEadulto_2023\dr_ESdEadulto_2023.xlsx")
# Path for the resulting json file
ESdE_MICRO_METADATA_PATH = Path(r"references\metadata\esde_adult_2023")


def read_raw_catalog(
    catalog:Path=RAW_ESdE_CATALOG_PATH
)->list[dict]:
    if not catalog.is_file():
        raise FileNotFoundError(f"Catalogo de variables no encontrado: {catalog}")
    
    # read_only=True would cause hyperlinks and merged cells to not work properly
    # data_only allows returning results rather than formulas in the cells.
    wb = load_workbook(filename = catalog, data_only=True) 

    sheet_diseño:worksheet = wb['Diseño']
    dictionary_list = extract_sheet_diseño(sheet_diseño)

    var_label_cache = {}
    # n=0
    for var_dict in dictionary_list:
        table_num_str = var_dict["Diccionario ubicado en la hoja…"]
        if table_num_str is None: continue

        # Original source column is not normalized
        table_num_norm:str=table_num_str.replace(" ", "")
        
        var_dict_name:str = var_dict["Diccionario de la variable"]

        cached_label = var_label_cache.get(var_dict_name)
        if cached_label is None:
            table_ws:worksheet = wb[table_num_norm]

            new_labels:dict = get_variable_options_from_table(table_ws, var_dict_name)

            var_label_cache[var_dict_name] = new_labels

            var_dict["value_labels"] = new_labels
            # logging.log(f"variable dictionary not cached: {labels}")

        else:
            labels = var_label_cache[var_dict_name]
            var_dict["value_labels"] = labels
            # logging.log(f"variable dictionary already cached: {labels}")
        # break

        # n+=1
        # if n >= 80:
        #     break

    return dictionary_list

def extract_sheet_diseño(
    wsheet:worksheet,
    max_row:int|None = 434
)->list[dict]:
    header_type_list = wsheet['A2':'J2'][0] # Select the only existing row

    dictionary_list:list[dict] = []
    variable_group:str|None = None

    for row in wsheet.iter_rows(min_row=3, max_row=max_row, min_col=1, max_col=12):
        row_map = {}

        for cell in row:
            if cell.column == 12:
                if cell.value is not None:
                    variable_group = cell.value

            elif cell.column <= len(header_type_list):
                value_type = header_type_list[cell.column - 1].value
                row_map[value_type] = cell.value

            # Col 11 stores URL about explanation of the statistic parameter... Not used for now in the project. 
            # elif cell.column == 11:
            #     row_map["URL_clasificacion"] = cell.value

        row_map["Grupo"] = variable_group
        dictionary_list.append(row_map)

    return dictionary_list

def get_variable_options_from_table(
    wsheet:worksheet,
    var_dict_name:str
)->dict:
    var_loc_cell = None
    for (cell,) in wsheet.iter_rows(min_row=5, max_col=1): # "(cell,)" because it removes unnecesary tuple encapsulation.
        if cell.value == var_dict_name:
            # logging.info(f"found variable label {cell.value} at {cell.coordinate}")
            var_loc_cell = cell
            break

    labels={}
    start_row = var_loc_cell.row + 2
    for code_cell, description_cell in wsheet.iter_rows(min_row=start_row, max_col=2):
        if code_cell.value is None and description_cell.value is None:
            # This might be brittle, if there is empty row between the labels of the same variable.
            # For now, it will suffice, as it seems there are no empty rows in 2023 dataset variable dictionary.
            break 

        # logging.info(f"({code_cell.value}) = {description_cell.value}")
        labels[str(code_cell.value)] = description_cell.value
        
    return labels


if __name__ == "__main__":
    logging.basicConfig(level=logging.INFO)
    dictionary_list = read_raw_catalog()

    n=0
    for map in dictionary_list:
        logging.info(map)
        n+=1
        if n >= 80:
            break