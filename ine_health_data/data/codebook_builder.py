'''
Generates a map keyed by variable code, where the structure for each variable is:
{
    'CCAA': {
        'Variable': 'CCAA', 
        'Diccionario de la variable': 'TCCAA', 
        'Longitud': 2, 
        'Tipo': 'A', 
        'Decimales': None, 
        'Posición': 1, 
        'Orden': 1, 
        'Diccionario ubicado en la hoja…': 'Tablas1', 
        'Descripción': 'Comunidad Autónoma de residencia', 
        'Observaciones': None, 
        'Grupo': 'DATOS DE IDENTIFICACIÓN ',
        'value_labels':{
            '01':'Andalucía',
            '02':'...',
        }
    }
}

This causes duplication of "value_labels" when multiple variables share the same dictionary.
Codebook contains 432 variables. Even with repeated number of value labels, it does not request the usage of chunking or
other optimization methods.

Positive trait of this structure is that it allows direct lookups by variable code and avoids lookups into multiple sources.
It could later be normalized by separating into different files for cross referencing or through insertion into SQL with duplication removal.

Possible error source (not in the 2023 microdata set) is duplication of variable codes. Should not happen, but if it were to,
it would overwrite the previous variable entry silently while building the map.
'''

from pathlib import Path
from openpyxl import load_workbook, worksheet
import logging

# Path for the original raw codebook supplied from INE
RAW_ESdE_CODEBOOK_PATH = Path(r"data\raw\datos_2023\ESdEadulto_2023\dr_ESdEadulto_2023.xlsx")
# Path for the resulting json file
CODEBOOK_OUTPUT_PATH = Path(r"references\metadata\esde_adult_2023")

# Seems that the codebook points to incorrect tables for multiple variables.
CODEBOOK_TABLE_OVERRIDES = {
    "T3H": "Tablas4",
    "T4H": "Tablas4",
}

def save_codebook():
    dictionary_list = extract_raw_codebook()

    from ..utils.json_helpers import save_json
    save_json(data=dictionary_list, output_path=CODEBOOK_OUTPUT_PATH)
    
    return

def extract_raw_codebook(
    codebook:Path=RAW_ESdE_CODEBOOK_PATH
)->list[dict]:
    if not codebook.is_file():
        raise FileNotFoundError(f"Catalogo de variables no encontrado: {codebook}")
    
    # read_only=True would cause hyperlinks and merged cells to not work properly
    # data_only allows returning results rather than formulas in the cells.
    wb = load_workbook(filename = codebook, data_only=True) 

    sheet_diseño:worksheet = wb['Diseño']
    dictionary_list = extract_sheet_diseño(sheet_diseño)

    var_label_cache = {}
    codebook_map = {}
    for var_dict in dictionary_list:
        table_num_str = var_dict["Diccionario ubicado en la hoja…"]
        if table_num_str is None: continue

        # Original source column is not normalized
        table_num_norm:str=table_num_str.replace(" ", "")
        
        var_dict_name:str = var_dict["Diccionario de la variable"]

        if var_dict_name in CODEBOOK_TABLE_OVERRIDES:
            table_num_norm = CODEBOOK_TABLE_OVERRIDES[var_dict_name]
            
        cached_label = var_label_cache.get(var_dict_name)
        if cached_label is None:
            table_ws:worksheet = wb[table_num_norm]

            new_labels:dict = get_variable_options_from_table(table_ws, var_dict_name)

            var_label_cache[var_dict_name] = new_labels

            var_dict["value_labels"] = new_labels
            # logging.debug(f"variable dictionary not cached: {labels}")

        else:
            labels = var_label_cache[var_dict_name]
            var_dict["value_labels"] = labels
            # logging.debug(f"variable dictionary already cached: {labels}")

        var_name = var_dict['Variable']
        codebook_map[var_name] = var_dict

    return codebook_map

def extract_sheet_diseño(
    wsheet:worksheet,
    max_row:int|None = 434
)->list[dict]:
    header_type_list = wsheet['A2':'J2'][0] # Select the only existing row

    dictionary_list:list[dict] = []
    variable_group:str|None = None

    for row in wsheet.iter_rows(min_row=3, max_row=max_row, min_col=1, max_col=12):
        row_map = {}

        for cell in row:
            if cell.column == 12:
                if cell.value is not None:
                    variable_group = cell.value

            elif cell.column <= len(header_type_list):
                value_type = header_type_list[cell.column - 1].value
                row_map[value_type] = cell.value

            # Col 11 stores URL about explanation of the statistic parameter... Not used for now in the project. 
            # elif cell.column == 11:
            #     row_map["URL_clasificacion"] = cell.value

        row_map["Grupo"] = variable_group
        dictionary_list.append(row_map)

    return dictionary_list

def get_variable_options_from_table(
    wsheet:worksheet,
    var_dict_name:str
)->dict:
    var_loc_cell = None
    for (cell,) in wsheet.iter_rows(min_row=5, max_col=1): # "(cell,)" because it removes unnecesary tuple encapsulation.
        if cell.value == var_dict_name:
            # logging.debug(f"found variable label {cell.value} at {cell.coordinate}")
            var_loc_cell = cell
            break

    if var_loc_cell is None: 
        raise RuntimeError(f"No variable options found for {var_dict_name} in {wsheet.title}.\n",
                           f"Check for possible source-reference errors, and add them to CODEBOOK_TABLE_OVERRIDES in metadata_builder.py")

    labels={}
    start_row = var_loc_cell.row + 2
    for code_cell, description_cell in wsheet.iter_rows(min_row=start_row, max_col=2):
        if code_cell.value is None and description_cell.value is None:
            # This might be brittle, if there is empty row between the labels of the same variable.
            # For now, it will suffice, as it seems there are no empty rows in 2023 dataset variable dictionary.
            break 

        # logging.info(f"({code_cell.value}) = {description_cell.value}")
        labels[str(code_cell.value)] = description_cell.value
        
    return labels

if __name__ == "__main__":
    logging.basicConfig(level=logging.DEBUG)
